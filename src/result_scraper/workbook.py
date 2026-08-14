"""Idempotent Excel persistence for scraped results."""

from __future__ import annotations

from datetime import date, datetime
import os
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

from .scraper import DrawResult, source_url, validate_prize


SHEET_NAME = "Nam Dinh Results"
HEADERS = (
    "Date",
    "Grand Prize Result",
    "1st Digit",
    "2nd Digit",
    "3rd Digit",
    "4th Digit",
    "5th Digit",
)
LEGACY_HEADERS = ("Date", "Grand Prize", "Variation C", "Variation D", None, None, None)
DEFAULT_OUTPUT = Path("outputs/result_scraping/nam_dinh_results.xlsx")


def _coerce_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return date.fromisoformat(value.strip())
    raise ValueError(f"invalid workbook date value: {value!r}")


def _new_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(HEADERS)
    return workbook


def load_existing_results(path: str | Path) -> dict[date, DrawResult]:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return {}

    workbook = load_workbook(workbook_path, data_only=False, read_only=True)
    try:
        sheet = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
        headers = tuple(sheet.cell(1, column).value for column in range(1, 8))
        if headers not in (HEADERS, LEGACY_HEADERS):
            raise ValueError(
                f"unexpected workbook headers {headers!r}; expected {HEADERS!r}"
            )

        results: dict[date, DrawResult] = {}
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            raw_date, raw_prize = row
            if raw_date in (None, "") and raw_prize in (None, ""):
                continue
            draw_date = _coerce_date(raw_date)
            prize = validate_prize(str(raw_prize).zfill(5))
            if draw_date in results:
                raise ValueError(f"duplicate date in workbook: {draw_date.isoformat()}")
            results[draw_date] = DrawResult(draw_date, prize, source_url(draw_date))
        return results
    finally:
        workbook.close()


def _style_sheet(sheet: object, last_row: int) -> None:
    navy = "17365D"
    pale_blue = "D9EAF7"
    white = "FFFFFF"
    thin_gray = Side(style="thin", color="D9E2F3")

    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:G{max(last_row, 1)}"
    sheet.print_title_rows = "1:1"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0

    sheet.column_dimensions["A"].width = 15
    sheet.column_dimensions["B"].width = 19
    for column in ("C", "D", "E", "F", "G"):
        sheet.column_dimensions[column].width = 13

    header_fill = PatternFill("solid", fgColor=navy)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(name="Aptos", size=11, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color=navy))
    sheet.row_dimensions[1].height = 24

    for row_index in range(2, last_row + 1):
        for column_index in range(1, 8):
            cell = sheet.cell(row_index, column_index)
            cell.font = Font(name="Aptos", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin_gray)
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=pale_blue)
        sheet.cell(row_index, 1).number_format = "yyyy-mm-dd"
        sheet.cell(row_index, 2).number_format = "00000"
        for column_index in range(3, 8):
            sheet.cell(row_index, column_index).number_format = "0"
        sheet.row_dimensions[row_index].height = 20


def upsert_results(path: str | Path, additions: Iterable[DrawResult]) -> int:
    """Insert or replace results by date and save atomically.

    Columns C through G are Excel formulas referencing Column B, keeping each
    digit extraction visible and auditable in the workbook.
    """

    output_path = Path(path)
    existing = load_existing_results(output_path)
    for result in additions:
        existing[result.draw_date] = result

    if output_path.exists():
        workbook = load_workbook(output_path, data_only=False)
        sheet = (
            workbook[SHEET_NAME]
            if SHEET_NAME in workbook.sheetnames
            else workbook.active
        )
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for table_name in list(sheet.tables):
            del sheet.tables[table_name]
    else:
        workbook = _new_workbook()
        sheet = workbook[SHEET_NAME]

    for column_index, header in enumerate(HEADERS, start=1):
        sheet.cell(1, column_index, header)

    for row_index, draw_date in enumerate(sorted(existing), start=2):
        result = existing[draw_date]
        sheet.cell(row_index, 1, draw_date)
        sheet.cell(row_index, 2, int(result.grand_prize))
        sheet.cell(row_index, 3, f"=INT(B{row_index}/10000)")
        sheet.cell(row_index, 4, f"=INT(MOD(B{row_index},10000)/1000)")
        sheet.cell(row_index, 5, f"=INT(MOD(B{row_index},1000)/100)")
        sheet.cell(row_index, 6, f"=INT(MOD(B{row_index},100)/10)")
        sheet.cell(row_index, 7, f"=MOD(B{row_index},10)")
        sheet.cell(row_index, 1).comment = Comment(
            f"Official source: {result.source_url}", "hachinokami789-maker"
        )

    last_row = len(existing) + 1
    _style_sheet(sheet, last_row)

    if last_row >= 2:
        table = Table(displayName="NamDinhResults", ref=f"A1:G{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)

    if workbook.calculation is None:
        workbook.calculation = CalcProperties()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.creator = "hachinokami789-maker"
    workbook.properties.title = "Nam Dinh Grand Prize Results"
    workbook.properties.subject = "Thinhnam daily lottery result history"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f"{output_path.stem}.tmp.xlsx")
    workbook.save(temporary_path)
    workbook.close()
    os.replace(temporary_path, output_path)
    return len(existing)


def ensure_workbook(path: str | Path) -> None:
    output_path = Path(path)
    if not output_path.exists():
        upsert_results(output_path, [])
